from classes import time_, geometry, Boundary 
import openpyxl

def conductivity_cylinder (str1, str2, str3, onedimens, R1, R2, times):

    with open (str1,'r') as f:
        initcond = float(f.read())
    with open (str2, "r") as f:
        coef = float(f.read())

    with open (str3, "w") as f:
        f.write("R1 boundary: \n")
        for i in range(len(R1.table)):
            f.write(str(R1.table[i][0]))
            f.write('\t')
            f.write(str(R1.table[i][1]))
            f.write('\n')
    with open (str3, "a") as f:
        f.write("R2 boundary: \n")
        for i in range(len(R2.table)):
            f.write(str(R2.table[i][0]) +'\t')
            f.write(str(R2.table[i][1]) +'\n')
        f.write('\n\n')
        f.write(f'Initial condition= {initcond} \n')
        f.write(f'Coefficient of thermal conductivity: {coef} \n' )
        f.write(f'Geometry: Raduis1 = {onedimens.radius1} \n Raduis2 = {onedimens.radius2} \n')

    flag = True 
    while flag:
        numb = int(input("Enter the number of steps in the space(by default enter 10): "))
        tau = float(input("Enter time step for the calculation (by default enter 0.5): "))
        h = ((onedimens.radius2 - onedimens.radius1) / numb) 
        if tau <= (h**2) / (2*coef): 
            flag = False 
        else:
            print("Error! Stability condition is not met")

    N1=len(R1.table)
    N2=len(R2.table)
    k = 0
    u=[None] * (numb+1)
    u_=[None] * (numb+1)
    alpha = [None] * (numb+1)
    beta = [None] * (numb+1)
    gamma = [None] * (numb+1)
        
    if R1.table[N1-1][0] > R2.table[N2-1][0]:
        times.duration = R1.table[N1-1][0]
    else:
        times.duration = R2.table[N2-1][0]
    times.set_step(tau)
    times.output = float(input("Enter time step for the output (by default enter 0.5): "))
    while times.output < times.step:
        times.output = float(input("Error! the step for output cannot be less than the step for calculation. Enter time step for the output (by default enter 0.5): "))
    times.set_n()
        
    u[0] = R1.table[0][1]
    u[numb] = R2.table[0][1]
    for i in range(1,numb):
            u[i] = initcond

    
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row = 1, column = 1).value = 't/x' 
    for i in range (numb+1):
        sheet.cell(row = 1, column = i+2).value = onedimens.radius1 + h * i

        
    sheet.cell(row = 2, column = 1).value = times.actual
    for i in range (numb+1):
        sheet.cell(row = 2, column = i+2).value = u[i]
           
    count = 1
    while times.actual < times.duration:

        if k == times.n:
            sheet.cell(row = count + 2, column = 1).value = times.actual
            for i in range (numb+1):
                sheet.cell(row = count+2, column = i+2).value = u[i]
            count += 1
            k = 0

        for i in range (numb+1):
            u_[i] = u[i]
        
        times.actual += times.step
        u[0] = R1.get(times.actual)
        u[numb] = R2.get(times.actual)

        for i in range (1,numb):
            A = ((-coef * tau) / ((h**2)*(onedimens.radius1 + h*(i-1)))) * ((onedimens.radius1 + h*(i-2) + onedimens.radius1 + h*(i-1)) / 2)
            B = 1 + (tau * coef/((h**2)*(onedimens.radius1 + h*(i-1)))) * ((onedimens.radius1 + h*(i-2) + 2*(onedimens.radius1 + h*(i-1)) + onedimens.radius1 + h*(i)) / 2)
            C = ((-coef * tau) / ((h**2)*(onedimens.radius1 + h*(i-1)))) * ((onedimens.radius1 + h*(i) + onedimens.radius1 + h*(i-1)) / 2)
            if i != numb-1 and i != 1:
                gamma[i] = B + A * alpha[i-1]
                beta[i] = (u_[i] - A*beta[i-1]) / gamma[i]
                alpha[i] = -C / gamma[i]
            elif i == 1:
                gamma [i] = B
                alpha[i] = -C / gamma[i]
                beta[i] = (u_[i] - A*u[0]) / gamma[i]
            else:
                gamma[i] = B + A * alpha[i-1]
                beta[i] = (u_[i] - C * u[numb] - A*beta[i-1]) / gamma[i]
                alpha[i] = 0
        
        u[numb-1] = beta[numb-1]
        for i in range (numb-2,0,-1):
            u[i] = alpha[i] * u[i+1] + beta[i]

        k += 1

    sheet.cell(row = count + 3, column = 1).value = times.actual
    for i in range(numb+1):
        sheet.cell(row = count+3, column = i+2).value = u[i]

    book.save("Solution of the equation cylinder.xlsx")
    book.close()
