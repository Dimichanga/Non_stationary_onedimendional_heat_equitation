from classes import time_, geometry, Boundary 
import openpyxl

def conductivity_flat (str1, str2, str3, onedimens = geometry, Left = Boundary, Right = Boundary, times = time_):

    with open (str1,'r') as f:
        initcond = float(f.read())
    with open (str2, "r") as f:
        coef = float(f.read())

    with open (str3, "w") as f:
        f.write("Left boundary: \n")
        for i in range(len(Left.table)):
            f.write(str(Left.table[i][0]))
            f.write('\t')
            f.write(str(Left.table[i][1]))
            f.write('\n')
    with open (str3, "a") as f:
        f.write('\n')
        for i in range(len(Left.table)):
            f.write(str(Right.table[i][0]) +'\t')
            f.write(str(Right.table[i][1]) +'\n')
        f.write('\n\n')
        f.write(f'Initial condition= {initcond} \n')
        f.write(f'Coefficient of thermal conductivity: {coef} \n' )
        f.write(f'Geometry: Length = {onedimens.length} \n')

    flag = True 
    while flag:
        numb = int(input("Enter the number of steps in the space(by default enter 10): "))
        tau = float(input("Enter time step for the calculation (by default enter 0.5): "))
        h = (onedimens.length / numb) 
        if tau <= (h**2) / (2*coef): 
            flag = False 
        else:
            print("Error! Stability condition is not met")

    Nl=len(Left.table)
    Nr=len(Right.table)
    k = 0
    u=[None] * (numb+1)
    u_=[None] * (numb+1)
    alpha = [None] * (numb+1)
    beta = [None] * (numb+1)
    gamma = [None] * (numb+1)
        
    if Left.table[Nl-1][0] > Right.table[Nr-1][0]:
        times.duration = Left.table[Nl-1][0]
    else:
        times.duration = Right.table[Nl-1][0]
    times.set_step(tau)
    times.output = float(input("Enter time step for the output (by default enter 0.5): "))
    while times.output < times.step:
        times.output = float(input("Error! the step for output cannot be less than the step for calculation. Enter time step for the output (by default enter 0.5): "))
    times.set_n()
        
    u[0] = Left.table[0][1]
    u[numb] = Right.table[0][1]
    for i in range(1,numb):
            u[i] = initcond

    
    book = openpyxl.Workbook()
    sheet = book.active
    sheet.cell(row = 1, column = 1).value = 't/x' 
    for i in range (numb+1):
        sheet.cell(row = 1, column = i+2).value = (onedimens.length / numb) * i

        
    sheet.cell(row = 2, column = 1).value = times.actual
    for i in range (numb+1):
        sheet.cell(row = 2, column = i+2).value = u[i]
        
       
    count = 1
    while times.actual < times.duration:

        if k == times.n:
            sheet.cell(row = count + 2, column = 1).value = times.actual
            for i in range (numb+1):
                sheet.cell(row = count+2, column = i+2).value = u[i]
            count += 1
            k = 0

        for i in range (numb+1):
            u_[i] = u[i]
        
        A = (-coef * tau) / (h**2)
        B = 1 + (2* tau * coef/(h**2))
        C = A
        times.actual += times.step
        u[0] = Left.get(times.actual)
        u[numb] = Right.get(times.actual)
        gamma [1] = B
        alpha[1] = -C / gamma[1]
        beta[1] = (u_[1] - A*u[0]) / gamma[1]

        for i in range (2,numb):
            gamma[i] = B + A * alpha[i-1]
            if i != numb-1:
                beta[i] = (u_[i] - A*beta[i-1]) / gamma[i]
                alpha[i] = -C / gamma[i]
            else:
                beta[i] = (u_[i] - A*u[numb] - A*beta[i-1]) / gamma[i]
                alpha[i] = 0
        
        u[numb-1] = beta[numb-1]
        for i in range (numb-2,0,-1):
            u[i] = alpha[i] * u[i+1] + beta[i]

        k += 1

    sheet.cell(row = count + 3, column = 1).value = times.actual
    for i in range(numb+1):
        sheet.cell(row = count+3, column = i+2).value = u[i]

    book.save("Solution of the equation.xlsx")
    book.close()